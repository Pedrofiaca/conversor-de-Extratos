import customtkinter as ctk
from tkinter import filedialog
import os
import re
import pandas as pd
import numpy as np
from docling.document_converter import DocumentConverter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
import os
import sys
import winreg


def encontrar_caminho_onedrive_registro():
    try:
        chave = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\OneDrive")
        caminho_onedrive, _ = winreg.QueryValueEx(chave, "UserFolder")
        winreg.CloseKey(chave)
        return caminho_onedrive
    except FileNotFoundError:
        return None
    
# Função para selecionar o arquivo PDF
def selecionar_pdf():
    arquivo_pdf = filedialog.askopenfilename(
        title="Selecione o PDF", 
        filetypes=(("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*"))
    )
    
    if arquivo_pdf:
        caminho_label.configure(text=arquivo_pdf)  # Exibir caminho do arquivo selecionado

# Função para iniciar o processo de conversão
def iniciar_conversao():
    caminho_pdf = caminho_label.cget("text")
    if caminho_pdf and caminho_pdf.endswith(".pdf"):
        aviso_label.configure(text="Processando o PDF...", text_color="green")
        processar_pdf(caminho_pdf)
    else:
        aviso_label.configure(text="Por favor, selecione um arquivo PDF válido.", text_color="red")

# Função para processar o PDF (integração com seu código)
def processar_pdf(caminho_pdf):
    # Converter o PDF em texto usando Docling
    converter = DocumentConverter()
    result = converter.convert(caminho_pdf)

    # Exportar o conteúdo convertido para Markdown
    markdown_content = result.document.export_to_markdown()

    # Salvar o conteúdo Markdown em um arquivo .md
    caminho_arquivo_md = 'extrato_bancario.md'  # Definindo o caminho do arquivo .md
    with open(caminho_arquivo_md, 'w', encoding='utf-8') as file:
        file.write(markdown_content)

    # Extração das tabelas e do nome
    tabelas_extraidas, nome_extraido = extrair_tabelas_md(caminho_arquivo_md)
    # Encontrar o caminho do OneDrive antes de criar a janela

    # Gerar o arquivo Excel com as tabelas extraídas e o nome da pessoa no nome do arquivo
    gerar_excel_com_tabelas(tabelas_extraidas, nome_extraido, f"{caminho_onedrive}\\Documentos\\Área de Trabalho\\Extratos Convertidos")
    aviso_label.configure(text="Processamento concluído! Arquivo Excel gerado.", text_color="green")

def extrair_nome(markdown_content):
    padrao_nome = r"Nome:\s*([A-Za-zÀ-ÿ\s]+)(?=\s*Extrato de:)"
    nome_match = re.search(padrao_nome, markdown_content)
    
    if nome_match:
        return nome_match.group(1).strip()
    return "Nome não encontrado"

def extrair_tabelas_md(caminho_arquivo_md):
    with open(caminho_arquivo_md, 'r', encoding='utf-8') as file:
        conteudo = file.read()
    
    padrao_tabela = r"(\|.+\|(?:\r?\n\|[-\s|:]+\|)+(?:\r?\n\|.+\|)+)"
    tabelas_md = re.findall(padrao_tabela, conteudo)
    
    tabelas = []
    nome = extrair_nome(conteudo)
    
    for tabela_md in tabelas_md:
        linhas = tabela_md.strip().split("\n")
        cabecalho = [celula.strip() for celula in linhas[0].split("|")[1:-1]]
        
        dados = []
        for linha in linhas[2:]:
            dados.append([celula.strip() for celula in linha.split("|")[1:-1]])
        
        df = pd.DataFrame(dados, columns=cabecalho)
        df = tratar_colunas(df)
        
        tabelas.append(df)
    
    return tabelas, nome

def tratar_colunas(df):
    df.rename(columns={
        'Data': 'Data da Transação',
        'Histórico': 'Descrição',
        'Docto.': 'Documento',
        'CrØdito (R$)': 'Crédito (R$)',
        'DØbito (R$)': 'Débito (R$)',
        'Saldo (R$)': 'Saldo Final'
    }, inplace=True)
    
    df['Descrição'] = df['Descrição'].apply(lambda x: re.split(r'\s{2,}', x)[0])
    
    df = df[~df['Data da Transação'].str.contains('Total', case=False, na=False)]
    
    df['Data da Transação'] = df['Data da Transação'].replace('', np.nan)
    df['Data da Transação'] = df['Data da Transação'].fillna(method='bfill')
    
    df['Crédito (R$)'].fillna(0, inplace=True)
    df['Débito (R$)'].fillna(0, inplace=True)
    
    return df

def gerar_excel_com_tabelas(tabelas, nome, caminho_arquivo_excel):
    df_final = pd.concat(tabelas, ignore_index=True)
    
    nome_arquivo = re.sub(r'[\\/*?:"<>|]', '', nome)
    nome_arquivo = nome_arquivo.replace("\n", " ").replace("\r", "").strip()
    
    if nome_arquivo.startswith("Extrato "):
        nome_arquivo = nome_arquivo.replace("Extrato ", "")
    
    nome_arquivo_final = f"Extrato {nome_arquivo}.xlsx"
    
    caminho_completo = f"{caminho_arquivo_excel}\\{nome_arquivo_final}"
    
    with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name="Transações")
        
        # Formatar como uma tabela
        workbook = writer.book
        worksheet = workbook["Transações"]
        
        table = Table(displayName="Tabela_Transacoes", ref=f"A1:{chr(64 + len(df_final.columns))}{len(df_final) + 1}")
        
        # Estilo da tabela
        style = TableStyleInfo(
            name="TableStyleLight8", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        table.tableStyleInfo = style
        
        worksheet.add_table(table)
        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF")  # Fonte branca no cabeçalho
    
    # Analisar as transações após a geração do Excel
    analisar_transacoes(caminho_completo)

def analisar_transacoes(caminho_arquivo_excel):
    # Carregar o Excel gerado
    df = pd.read_excel(caminho_arquivo_excel, sheet_name="Transações")
    
    # Contar quantas vezes cada transação aparece na coluna 'Descrição'
    transacao_contagem = df['Descrição'].value_counts().reset_index()
    transacao_contagem.columns = ['Descrição', 'Contagem']
    
    # Carregar o arquivo Excel existente para adicionar uma nova aba
    with pd.ExcelWriter(caminho_arquivo_excel, engine='openpyxl', mode='a') as writer:
        # Carregar o workbook existente
        workbook = writer.book
        # Adicionar a contagem como uma nova aba
        transacao_contagem.to_excel(writer, sheet_name='Contagem de Transações', index=False)
        
        # Acessar a nova aba e formatar como uma tabela
        worksheet = workbook['Contagem de Transações']
        table = Table(displayName="Tabela_Contagem", ref=f"A1:B{len(transacao_contagem) + 1}")
        
        # Adicionar estilo de tabela
        style = TableStyleInfo(
            name="TableStyleLight8", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        table.tableStyleInfo = style
        
        worksheet.add_table(table)
        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF")  # Fonte branca no cabeçalho
    
    print(f"Contagem das transações adicionada como tabela no arquivo: {caminho_arquivo_excel}")

    # Encontrar o caminho do OneDrive antes de criar a janela
caminho_onedrive = encontrar_caminho_onedrive_registro()

# Configurações iniciais da janela principal
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

root = ctk.CTk()
root.title("Coversor de Extratos - M&J Advogados")
root.geometry("500x250")  # Ajustando altura para acomodar novos elementos

# Caminho absoluto para o ícone
icon_path = os.path.join(os.path.dirname(__file__), "Logo_v2.ico")
root.iconbitmap(icon_path)

# Layout da interface gráfica
frame = ctk.CTkFrame(root)
frame.pack(pady=20, padx=40, fill="both", expand=True)

# Botão para selecionar o arquivo PDF
selecao_btn = ctk.CTkButton(frame, text="Selecionar PDF", command=selecionar_pdf)
selecao_btn.pack(pady=10)

# Label para mostrar o caminho do arquivo selecionado
caminho_label = ctk.CTkLabel(frame, text="Nenhum arquivo selecionado", wraplength=400)
caminho_label.pack(pady=10)

# Aviso ou status do processamento
aviso_label = ctk.CTkLabel(frame, text="", wraplength=400)
aviso_label.pack(pady=10)

# Reposicionando o botão "Iniciar Conversão"
iniciar_btn = ctk.CTkButton(frame, text="Gerar Excel", command=iniciar_conversao)
iniciar_btn.pack(pady=10)
# Executar a interface
root.mainloop()
