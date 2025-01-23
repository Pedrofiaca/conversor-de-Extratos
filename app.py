from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import os
from docling.document_converter import DocumentConverter
import pandas as pd
import re
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)

# Diretórios para upload e saída
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        
        if file and file.filename.endswith('.pdf'):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)
            result_path = processar_pdf(file_path)  # Chama o processo de conversão
            return redirect(url_for('resultado', filename=os.path.basename(result_path)))
        
        return redirect(request.url)

    # Se for GET, exibe o formulário de upload
    return render_template('upload.html')

@app.route('/resultado/<filename>')
def resultado(filename):
    return render_template('resultado.html', filename=filename)

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if not os.path.exists(file_path):
        print(f"Erro: arquivo não encontrado em {file_path}")
        return "Arquivo não encontrado", 404
    return send_from_directory(app.config['OUTPUT_FOLDER'], filename)
    
def processar_pdf(caminho_pdf):
    # Processamento do PDF
    converter = DocumentConverter()
    result = converter.convert(caminho_pdf)
    markdown_content = result.document.export_to_markdown()

    tabelas_extraidas, nome_extraido = extrair_tabelas_md(markdown_content)
    caminho_excel = os.path.join(app.config['OUTPUT_FOLDER'], f"Extrato_{nome_extraido}.xlsx")
    gerar_excel_com_tabelas(tabelas_extraidas, nome_extraido, caminho_excel)
    return caminho_excel

def extrair_tabelas_md(markdown_content):
    padrao_tabela = r"(\|.+\|(?:\r?\n\|[-\s|:]+\|)+(?:\r?\n\|.+\|)+)"
    tabelas_md = re.findall(padrao_tabela, markdown_content)
    tabelas = []
    nome = extrair_nome(markdown_content)

    for tabela_md in tabelas_md:
        linhas = tabela_md.strip().split("\n")
        cabecalho = [celula.strip() for celula in linhas[0].split("|")[1:-1]]
        dados = [[celula.strip() for celula in linha.split("|")[1:-1]] for linha in linhas[2:]]
        df = pd.DataFrame(dados, columns=cabecalho)
        df = tratar_colunas(df)
        tabelas.append(df)

    return tabelas, nome

def tratar_colunas(df):
    df.rename(columns={
        'Data': 'Data da Transação',
        'Histórico': 'Descrição',
        'Docto.': 'Documento',
        'CrØdito (R$)': 'Crédito (R$)',
        'DØbito (R$)': 'Débito (R$)',
        'Saldo (R$)': 'Saldo Final'
    }, inplace=True)
    
    df['Descrição'] = df['Descrição'].apply(lambda x: re.split(r'\s{2,}', x)[0])
    
    df = df[~df['Data da Transação'].str.contains('Total', case=False, na=False)]
    
    df['Data da Transação'] = df['Data da Transação'].replace('', np.nan)
    df['Data da Transação'] = df['Data da Transação'].fillna(method='bfill')
    
    df['Crédito (R$)'].fillna(0, inplace=True)
    df['Débito (R$)'].fillna(0, inplace=True)
    
    return df

def extrair_nome(markdown_content):
    padrao_nome = r"Nome:\s*([A-Za-zÀ-ÿ\s]+)(?=\s*Extrato de:)"
    nome_match = re.search(padrao_nome, markdown_content)
    return nome_match.group(1).strip() if nome_match else "Desconhecido"

def gerar_excel_com_tabelas(tabelas, nome, caminho_arquivo_excel):
    df_final = pd.concat(tabelas, ignore_index=True)
    with pd.ExcelWriter(caminho_arquivo_excel, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name="Transações")
        # Formatar como uma tabela
        workbook = writer.book
        worksheet = workbook["Transações"]
        
        table = Table(displayName="Tabela_Transacoes", ref=f"A1:{chr(64 + len(df_final.columns))}{len(df_final) + 1}")
        
        # Estilo da tabela
        style = TableStyleInfo(
            name="TableStyleLight8", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        table.tableStyleInfo = style
        
        worksheet.add_table(table)
        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF")  # Fonte branca no cabeçalho

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Letra da coluna
        for cell in column_cells:
            try:
                if cell.value:  # Verifica se o valor não é None
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
    adjusted_width = max_length + 2  # Adiciona um espaçamento
    worksheet.column_dimensions[column].width = adjusted_width
    
if __name__ == '__main__':
    app.run(debug=True)
